"""
Sala de Compresores — Calculadora de dimensionamiento
Autor: Daniel Ignacio Auger Solis  |  daniel.auger1998@gmail.com
"""
import math
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ══════════════════════════════════════════════════════════════
#  CONSTANTES FÍSICAS
# ══════════════════════════════════════════════════════════════
T_N = 293.15   # 20°C
P_N = 1.01325     # bar abs

delta_p_coefs = {
    0.10: (0.04668, 0.7647),
    0.16: (0.07018, 0.7656),
    0.25: (0.1201,  0.7639),
    0.40: (0.1866,  0.7726),
    0.63: (0.313,   0.7581),
    1.00: (0.4651,  0.7652),
    1.60: (0.817,   0.7553),
}
z_coefs = {
    10: (0.7097, 0.7657),
    15: (0.4657, 0.7684),
    20: (0.2692, 0.8056),
}

# ══════════════════════════════════════════════════════════════
#  PALETA DE COLORES
# ══════════════════════════════════════════════════════════════
C = {
    "bg":      "#0d0d0d",
    "panel":   "#141414",
    "card":    "#1a1a1a",
    "border":  "#2a2a2a",
    "border2": "#3a3a3a",
    "text":    "#d8d8d8",
    "text2":   "#888888",
    "text3":   "#444444",
    "accent":  "#00bfff",
    "accent2": "#00e5a0",
    "accent3": "#ffa040",
    "green":   "#00e5a0",
    "red":     "#ff4455",
    "entry":   "#0f0f0f",
    "sel":     "#1e3a4a",
    "hdr":     "#080808",
    "ok":      "#00c853",
    "warn":    "#ffaa00",
}

# ══════════════════════════════════════════════════════════════
#  FUNCIONES DE CÁLCULO
# ══════════════════════════════════════════════════════════════
def fmt(v, d=4):
    """Número con coma decimal."""
    return f"{v:.{d}f}".replace(".", ",")

def parse(s):
    return float(str(s).replace(",", "."))

def a_Nl_min(q, tipo, p_ref_bg, t_ref_c, p_atm_abs, t_amb_c=20.0):
    if tipo == "FAD (m³/min)":
        return q * 1000.0

    elif tipo == "scfm":
        return q * 26.3526

    elif tipo == "Normalizado (Nl/min)":
        return q

    elif tipo == "Personalizado":
        T_ref = (t_ref_c if t_ref_c is not None else 0.0) + 273.15
        P_ref = p_ref_bg + p_atm_abs    # bar g → bar abs
        # q_N [Nl/min] = q [m³/min] × (P_ref/P_N) × (T_N/T_ref) × 1000
        return q * (P_ref / P_N) * (T_N / T_ref) * 1000.0

    return 0.0

def Nl_min_a_FAD(qN, t_amb_c, p_atm_abs):
    T_amb = t_amb_c + 273.15
    return qN * (P_N / p_atm_abs) * (T_amb / T_N) / 1000.0

def calc_deposito(q_l, dp_val, z_val):
    a_dp, b_dp = delta_p_coefs[dp_val]
    a_z,  b_z  = z_coefs[z_val]
    if q_l <= a_dp:
        return None
    x = math.log(q_l / a_dp) / b_dp
    return a_z * math.exp(b_z * x)

# ══════════════════════════════════════════════════════════════
#  HELPERS DE WIDGETS
# ══════════════════════════════════════════════════════════════
TIPOS = ["FAD (m³/min)", "scfm", "Normalizado (Nl/min)", "Personalizado"]

# Anchos de columna en caracteres (Entry y encabezado comparten)
COL_DEFS = [
    ("nombre", 15, "Equipo"),
    ("cant",    5, "Cant."),
    ("p_req",   8, "P req (bar g)"),
    ("caudal",  9, "Caudal"),
    ("tipo",   19, "Tipo caudal"),
    ("uso",     6, "Uso (%)"),
    ("p_ref",   8, "P ref (bar g)"),
    ("t_ref",   7, "T ref (°C)"),
    ("del",     3, ""),
]
# Padding horizontal por columna en px (igual para hdr y fila)
COL_PAD = 3

def mk_entry(parent, textvariable=None, width=10, state="normal"):
    return tk.Entry(
        parent,
        textvariable=textvariable, width=width, state=state,
        bg=C["entry"], fg=C["text"],
        insertbackground=C["accent"],
        disabledbackground=C["bg"],
        disabledforeground=C["text3"],
        relief="flat", bd=0,
        highlightthickness=1,
        highlightbackground=C["border2"],
        highlightcolor=C["accent"],
        font=("Consolas", 10),
    )

def mk_lbl(parent, text, fg=None, font=None, **kw):
    bg = C["panel"]
    try:
        bg = parent.cget("bg")
    except Exception:
        pass
    return tk.Label(parent, text=text, bg=bg,
                    fg=fg or C["text2"],
                    font=font or ("Segoe UI", 10), **kw)

def mk_radio(parent, text, variable, value, command=None):
    bg = C["panel"]
    try:
        bg = parent.cget("bg")
    except Exception:
        pass
    return tk.Radiobutton(
        parent, text=text, variable=variable, value=value,
        command=command,
        bg=bg, fg=C["text"],
        activebackground=bg, activeforeground=C["accent"],
        selectcolor=C["border"],
        font=("Segoe UI", 10),
    )

def mk_sep(parent):
    tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", pady=5)

# ══════════════════════════════════════════════════════════════
#  FILA DE EQUIPO
# ══════════════════════════════════════════════════════════════
class FilaEquipo:
    def __init__(self, parent, on_delete):
        self._on_delete_cb = on_delete
        self.vars = {k: tk.StringVar() for k, *_ in COL_DEFS}

        # self.row vive en frame_tabla; _agregar_fila asigna el grid row
        self.row = tk.Frame(parent, bg=C["panel"])

        self._widgets = {}
        for col_idx, (key, cw, _) in enumerate(COL_DEFS):
            if key == "tipo":
                w = ttk.Combobox(self.row, textvariable=self.vars[key],
                                 values=TIPOS, state="readonly",
                                 width=cw - 2, font=("Consolas", 10))
                w.bind("<<ComboboxSelected>>", self._on_tipo)
            elif key == "del":
                w = tk.Button(self.row, text="✕",
                              bg=C["panel"], fg=C["red"],
                              activebackground=C["red"], activeforeground="white",
                              relief="flat", bd=0, padx=4, cursor="hand2",
                              font=("Consolas", 10, "bold"),
                              command=lambda: self._on_delete_cb(self))
            else:
                w = mk_entry(self.row, self.vars[key], cw)
            w.grid(row=0, column=col_idx, padx=COL_PAD, pady=2, sticky="w")
            self._widgets[key] = w

        self._on_tipo()

    def _on_tipo(self, *_):
        custom = self.vars["tipo"].get() == "Personalizado"
        for k in ("p_ref", "t_ref"):
            self._widgets[k].config(
                state="normal" if custom else "disabled",
                highlightbackground=C["border2"] if custom else C["border"])

    def obtener_datos(self):
        nombre = self.vars["nombre"].get().strip() or "Equipo"
        cant   = int(self.vars["cant"].get() or "1")
        p_req  = parse(self.vars["p_req"].get() or "0")
        caudal = parse(self.vars["caudal"].get() or "0")
        tipo   = self.vars["tipo"].get()
        uso    = parse(self.vars["uso"].get() or "100")
        p_ref = t_ref = None
        if tipo == "Personalizado":
            s = self.vars["p_ref"].get().strip()
            if not s:
                raise ValueError(f'"{nombre}": Personalizado requiere Presión de referencia.')
            p_ref = parse(s)
            ts = self.vars["t_ref"].get().strip()
            t_ref = parse(ts) if ts else 0.0
        return dict(nombre=nombre, cant=cant, p_req=p_req, caudal=caudal,
                    tipo=tipo, uso=uso, p_ref=p_ref, t_ref=t_ref)

    def to_dict(self):
        return {k: v.get() for k, v in self.vars.items() if k != "del"}

    def from_dict(self, d):
        for k, v in d.items():
            if k in self.vars:
                self.vars[k].set(v)
        self.vars["tipo"].set(d.get("tipo", ""))
        self._widgets["tipo"].set(d.get("tipo", ""))
        self._on_tipo()

    def destroy(self):
        self.row.grid_forget()
        self.row.destroy()

# ══════════════════════════════════════════════════════════════
#  DIAGRAMA CANVAS
# ══════════════════════════════════════════════════════════════
class Diagrama(tk.Canvas):
    """Canvas que dibuja compresor → tubería → depósito con valores encima."""
    MARGIN_TOP = 52   # px reservados para textos sobre los bloques

    def __init__(self, parent, **kw):
        super().__init__(parent, bg=C["card"],
                         highlightthickness=1,
                         highlightbackground=C["border2"],
                         height=240, **kw)
        self.q_val = "—"
        self.p_val = "—"
        self.v_val = "—"
        self.bind("<Configure>", lambda e: self._draw())

    def update_values(self, q, p, v):
        self.q_val, self.p_val, self.v_val = q, p, v
        self._draw()

    def _draw(self):
        self.delete("all")
        W = self.winfo_width()
        H = self.winfo_height()
        if W < 20:
            return

        self.create_rectangle(0, 0, W, H, fill=C["card"], outline="")

        MT = self.MARGIN_TOP    # top margin for labels
        MB = 22                 # bottom margin for name labels

        cw, ch = 130, 80
        dw, dh = 90, 95
        gap    = 80
        total  = cw + gap + dw
        x0     = (W - total) // 2
        # vertically center the blocks in remaining space
        avail  = H - MT - MB
        ym     = MT + avail // 2

        # ── COMPRESOR ──────────────────────────────
        cx1, cy1 = x0,      ym - ch // 2
        cx2, cy2 = x0 + cw, ym + ch // 2

        self.create_rectangle(cx1+3, cy1+3, cx2+3, cy2+3, fill="#060606", outline="")
        self.create_rectangle(cx1, cy1, cx2, cy2,
                              fill="#152030", outline=C["accent"], width=2)
        for i in (1, 2, 3):
            yy = cy1 + ch * i // 4
            self.create_line(cx1+6, yy, cx2-6, yy, fill="#1a3040", width=1)

        gcx, gcy = (cx1+cx2)//2, (cy1+cy2)//2
        self.create_oval(gcx-22, gcy-22, gcx+22, gcy+22,
                         fill="#0a1c2a", outline=C["accent"], width=2)
        self.create_oval(gcx-10, gcy-10, gcx+10, gcy+10,
                         fill=C["accent"], outline="")
        for ang in range(0, 360, 45):
            r = math.radians(ang)
            self.create_line(gcx+13*math.cos(r), gcy+13*math.sin(r),
                             gcx+25*math.cos(r), gcy+25*math.sin(r),
                             fill=C["accent"], width=3, capstyle="round")

        # Etiqueta COMPRESOR — debajo del bloque con margen
        self.create_text((cx1+cx2)//2, cy2 + 12,
                         text="COMPRESOR",
                         fill=C["text3"], font=("Consolas", 8, "bold"))

        # Valores Q y P — ARRIBA del bloque, dentro del margen reservado
        vx = (cx1+cx2)//2
        self.create_text(vx, cy1 - 28,
                         text=f"Q = {self.q_val} m³/min",
                         fill=C["accent"], font=("Consolas", 9, "bold"))
        self.create_text(vx, cy1 - 12,
                         text=f"P = {self.p_val} bar g",
                         fill=C["accent2"], font=("Consolas", 9, "bold"))

        # ── TUBERÍA ────────────────────────────────
        pipe_y = ym
        px1, px2 = cx2, x0 + cw + gap
        self.create_line(px1, pipe_y, px2, pipe_y,
                         fill=C["accent"], width=5, capstyle="round")
        for fx in range(px1+16, px2-8, 20):
            self.create_polygon(fx, pipe_y-6, fx+10, pipe_y, fx, pipe_y+6,
                                fill=C["accent2"], outline="")
        for bx in (px1, px2):
            self.create_rectangle(bx-3, pipe_y-9, bx+3, pipe_y+9,
                                  fill=C["border2"], outline=C["accent"], width=1)

        # ── DEPÓSITO ───────────────────────────────
        dx1 = x0 + cw + gap
        dy1 = ym - dh // 2
        dx2 = dx1 + dw
        dy2 = dy1 + dh
        er  = dw // 2

        self.create_rectangle(dx1+3, dy1+3, dx2+3, dy2+3, fill="#060606", outline="")
        self.create_rectangle(dx1, dy1, dx2, dy2,
                              fill="#102018", outline=C["accent2"], width=2)
        fill_h = int((dh - 8) * 0.60)
        self.create_rectangle(dx1+3, dy2-fill_h, dx2-3, dy2-4,
                              fill="#0a1810", outline="")
        for i in (1, 2, 3):
            ly = dy2 - fill_h + fill_h*i//4
            self.create_line(dx1+5, ly, dx2-5, ly, fill="#1a2a1a", width=1)
        # tapas elípticas
        for ey in (dy1, dy2):
            self.create_oval(dx1, ey - er//3, dx2, ey + er//3,
                             fill="#142018", outline=C["accent2"], width=2)

        # Etiqueta DEPÓSITO — debajo con margen suficiente
        self.create_text((dx1+dx2)//2, dy2 + 14,
                         text="DEPÓSITO",
                         fill=C["text3"], font=("Consolas", 8, "bold"))

        # Valor V — bien arriba, dentro del margen reservado
        self.create_text((dx1+dx2)//2, dy1 - 14,
                         text=f"V = {self.v_val} m³",
                         fill=C["accent3"], font=("Consolas", 9, "bold"))


# ══════════════════════════════════════════════════════════════
#  PANEL COLAPSABLE CON INDICADOR DE ESTADO
# ══════════════════════════════════════════════════════════════
class SeccionPanel:
    """
    Sección con botón de apertura/cierre y pastilla de estado:
      estado="vacio"    → pastilla gris
      estado="ok"       → pastilla verde  ✔
      estado="error"    → pastilla roja   ✖
    """
    def __init__(self, parent, numero, titulo, color_acento=None):
        self.color = color_acento or C["accent"]
        self._open = False

        outer = tk.Frame(parent, bg=C["border2"],
                         highlightthickness=0)
        outer.pack(fill="x", padx=18, pady=4)

        # ── Barra del botón ──
        self.btn_bar = tk.Frame(outer, bg=C["border"], height=34, cursor="hand2")
        self.btn_bar.pack(fill="x")
        self.btn_bar.pack_propagate(False)
        self.btn_bar.bind("<Button-1>", self._toggle)

        # Número
        tk.Label(self.btn_bar, text=f" {numero}",
                 bg=C["border"], fg=self.color,
                 font=("Consolas", 10, "bold")).pack(side="left", padx=(8, 0))

        # Triángulo
        self.lbl_tri = tk.Label(self.btn_bar, text="▶",
                                bg=C["border"], fg=C["text3"],
                                font=("Consolas", 9))
        self.lbl_tri.pack(side="left", padx=4)
        self.lbl_tri.bind("<Button-1>", self._toggle)

        # Título
        self.lbl_tit = tk.Label(self.btn_bar,
                                text=titulo.upper(),
                                bg=C["border"], fg=self.color,
                                font=("Consolas", 9, "bold"))
        self.lbl_tit.pack(side="left", padx=4)
        self.lbl_tit.bind("<Button-1>", self._toggle)

        # Pastilla estado
        self.lbl_estado = tk.Label(self.btn_bar, text="  ●  SIN DATOS  ",
                                   bg=C["text3"], fg=C["bg"],
                                   font=("Consolas", 8, "bold"),
                                   padx=6, pady=2)
        self.lbl_estado.pack(side="right", padx=10)

        # ── Cuerpo colapsable ──
        self.body = tk.Frame(outer, bg=C["panel"], padx=14, pady=10)
        # No se empaqueta hasta que se abra

    def _toggle(self, *_):
        if self._open:
            self.body.pack_forget()
            self.lbl_tri.config(text="▶")
        else:
            self.body.pack(fill="x")
            self.lbl_tri.config(text="▼")
        self._open = not self._open

    def abrir(self):
        if not self._open:
            self._toggle()

    def cerrar(self):
        if self._open:
            self._toggle()

    def set_estado(self, estado, texto=""):
        if estado == "ok":
            self.lbl_estado.config(bg=C["ok"],   fg=C["bg"],
                                   text=f"  ✔  {texto or 'LISTO'}  ")
        elif estado == "error":
            self.lbl_estado.config(bg=C["red"],  fg="white",
                                   text=f"  ✖  {texto or 'INCOMPLETO'}  ")
        else:
            self.lbl_estado.config(bg=C["text3"], fg=C["bg"],
                                   text="  ●  SIN DATOS  ")


# ══════════════════════════════════════════════════════════════
#  APLICACIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════
class App:
    def __init__(self, root):
        self.root  = root
        root.title("SALA DE COMPRESORES")
        root.geometry("1280x960")
        root.configure(bg=C["bg"])
        root.resizable(True, True)
        self.filas = []
        self._apply_styles()
        self._build()

    # ─── ttk styles ────────────────────────────
    def _apply_styles(self):
        s = ttk.Style(self.root)
        try:
            s.theme_use("clam")
        except Exception:
            pass
        s.configure(".", background=C["card"], foreground=C["text"],
                    fieldbackground=C["entry"], troughcolor=C["border"],
                    selectbackground=C["sel"], selectforeground=C["accent"])
        s.configure("TCombobox",
                    background=C["entry"], foreground=C["text"],
                    fieldbackground=C["entry"], arrowcolor=C["accent"],
                    font=("Consolas", 10))
        s.map("TCombobox",
              fieldbackground=[("readonly", C["entry"])],
              foreground=[("readonly", C["text"])],
              selectbackground=[("readonly", C["sel"])])
        s.configure("TScrollbar",
                    background=C["border"], troughcolor=C["bg"],
                    arrowcolor=C["text2"])

    # ─── Estructura principal ───────────────────
    def _build(self):
        self._menubar()
        self._titlebar()

        outer = tk.Frame(self.root, bg=C["bg"])
        outer.pack(fill="both", expand=True)

        self._canvas = tk.Canvas(outer, bg=C["bg"],
                                 highlightthickness=0, bd=0)
        vsb = ttk.Scrollbar(outer, orient="vertical",
                            command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self.main = tk.Frame(self._canvas, bg=C["bg"])
        wid = self._canvas.create_window((0, 0), window=self.main, anchor="nw")

        self.main.bind("<Configure>",
            lambda e: self._canvas.configure(
                scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>",
            lambda e: self._canvas.itemconfig(wid, width=e.width))
        self._canvas.bind_all("<MouseWheel>",
            lambda e: self._canvas.yview_scroll(
                int(-1*(e.delta/120)), "units"))

        self._sec1_ambient()
        self._sec2_servicios()
        self._sec3_compresor()
        self._sec4_deposito()
        self._boton_calcular()
        self._sec5_resultados()

    # ─── Menú superior ─────────────────────────
    def _menubar(self):
        bar = tk.Frame(self.root, bg="#050505", height=30)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        def btn(txt, cmd):
            b = tk.Button(bar, text=txt, command=cmd,
                          bg="#050505", fg=C["text2"],
                          activebackground=C["border"],
                          activeforeground=C["accent"],
                          relief="flat", bd=0, padx=12, pady=0,
                          font=("Consolas", 9), cursor="hand2")
            b.pack(side="left")
            return b

        btn("💾  Guardar proyecto", self._guardar)
        btn("📂  Cargar proyecto",  self._cargar)
        btn("📊  Exportar a Excel", self._exportar_excel)

        # Acerca de (derecha)
        info = tk.Label(
            bar,
            text="Daniel Ignacio Auger Solis  ·  daniel.auger1998@gmail.com",
            bg="#050505", fg=C["text3"],
            font=("Consolas", 8))
        info.pack(side="right", padx=12)

        about_btn = tk.Button(
            bar, text="ℹ  Acerca de",
            command=self._about,
            bg="#050505", fg=C["text2"],
            activebackground=C["border"],
            activeforeground=C["accent"],
            relief="flat", bd=0, padx=12,
            font=("Consolas", 9), cursor="hand2")
        about_btn.pack(side="right")

        tk.Frame(self.root, bg=C["border"], height=1).pack(fill="x")

    def _about(self):
        win = tk.Toplevel(self.root)
        win.title("Acerca de")
        win.configure(bg=C["card"])
        win.geometry("440x220")
        win.resizable(False, False)
        tk.Label(win, text="SALA DE COMPRESORES",
                 bg=C["card"], fg=C["accent"],
                 font=("Consolas", 13, "bold")).pack(pady=(24, 6))
        tk.Label(win, text="Calculadora de dimensionamiento de compresor y depósito",
                 bg=C["card"], fg=C["text2"],
                 font=("Segoe UI", 10)).pack()
        tk.Frame(win, bg=C["border"], height=1).pack(fill="x", pady=16, padx=30)
        tk.Label(win, text="Daniel Ignacio Auger Solis",
                 bg=C["card"], fg=C["text"],
                 font=("Segoe UI", 10, "bold")).pack()
        tk.Label(win, text="daniel.auger1998@gmail.com",
                 bg=C["card"], fg=C["accent2"],
                 font=("Consolas", 10)).pack(pady=4)
        tk.Button(win, text="Cerrar", command=win.destroy,
                  bg=C["border"], fg=C["text"],
                  relief="flat", bd=0, padx=20, pady=6,
                  font=("Consolas", 9), cursor="hand2").pack(pady=14)

    # ─── Barra de título ───────────────────────
    def _titlebar(self):
        bar = tk.Frame(self.root, bg=C["hdr"], height=48)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Label(bar, text="SALA DE COMPRESORES",
                 bg=C["hdr"], fg=C["accent"],
                 font=("Consolas", 13, "bold")).pack(side="left", padx=18, pady=12)
        tk.Label(bar, text="dimensionamiento de compresor y depósito",
                 bg=C["hdr"], fg=C["text3"],
                 font=("Consolas", 9)).pack(side="left", pady=14)
        tk.Frame(self.root, bg=C["accent"], height=2).pack(fill="x")

    # ─── Helpers de layout ─────────────────────
    def _card(self, title, color=None):
        outer = tk.Frame(self.main, bg=C["border2"])
        outer.pack(fill="x", padx=18, pady=4)
        hdr = tk.Frame(outer, bg=C["border"], height=30)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=title.upper(),
                 bg=C["border"], fg=color or C["accent"],
                 font=("Consolas", 9, "bold"),
                 padx=12).pack(side="left", fill="y")
        body = tk.Frame(outer, bg=C["panel"], padx=14, pady=10)
        body.pack(fill="x")
        return body

    # ══════════════════════════════════════════
    #  SECCIÓN 1 – CONDICIONES AMBIENTALES
    # ══════════════════════════════════════════
    def _sec1_ambient(self):
        self._sec1 = SeccionPanel(self.main, "1", "Condiciones ambientales")
        body = self._sec1.body

        r = tk.Frame(body, bg=C["panel"])
        r.pack(fill="x")
        mk_lbl(r, "Temperatura máxima ambiente (°C)").pack(side="left", padx=(0,6))
        self.var_t_amb = tk.StringVar()
        mk_entry(r, self.var_t_amb, 8).pack(side="left", padx=(0,28))
        mk_lbl(r, "Presión atmosférica local (bar abs)").pack(side="left", padx=(0,6))
        self.var_p_atm = tk.StringVar()
        mk_entry(r, self.var_p_atm, 11).pack(side="left")

        mk_sep(body)
        tk.Button(body, text="✔  Confirmar condiciones ambientales",
                  bg=C["ok"], fg=C["bg"],
                  activebackground="#00a040", activeforeground=C["bg"],
                  relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
                  font=("Consolas", 9, "bold"),
                  command=self._confirmar_sec1).pack(anchor="w")

    def _confirmar_sec1(self):
        t = self.var_t_amb.get().strip()
        p = self.var_p_atm.get().strip()
        if not t or not p:
            self._sec1.set_estado("error", "FALTAN DATOS")
            return
        try:
            parse(t); parse(p)
        except ValueError:
            self._sec1.set_estado("error", "VALORES INVÁLIDOS")
            return
        self._sec1.set_estado("ok", f"T={fmt(parse(t),1)}°C  P={fmt(parse(p),5)} bar")
        self._sec1.cerrar()

    # ══════════════════════════════════════════
    #  SECCIÓN 2 – SERVICIOS
    # ══════════════════════════════════════════
    def _sec2_servicios(self):
        self._sec2 = SeccionPanel(self.main, "2",
                                  "Servicios que requieren aire comprimido")
        body = self._sec2.body

        # ── Tabla con alineación perfecta ─────────────────────────────────
        # frame_tabla es el único grid-master. Fila 0 = header (Labels),
        # filas 1..N = una FilaEquipo cada una.
        # Todos los widgets de header y de fila comparten la misma columna
        # de grid, por lo que tkinter garantiza alineación exacta.
        self.frame_tabla = tk.Frame(body, bg=C["panel"])
        self.frame_tabla.pack(fill="x", anchor="w")
        self._next_eq_row = 1   # fila 0 reservada para el header

        # Header: Labels en fila 0 del mismo frame_tabla
        for col_idx, (key, cw, title) in enumerate(COL_DEFS):
            tk.Label(self.frame_tabla, text=title,
                     bg=C["border"], fg=C["accent"],
                     font=("Consolas", 10, "bold"),
                     anchor="w").grid(row=0, column=col_idx,
                                      padx=COL_PAD, pady=(3,2), sticky="w")

        # frame_eq es alias para que _agregar_fila sepa dónde crear las filas
        self.frame_eq = self.frame_tabla

        mk_lbl(body,
               "P ref y T ref: solo para 'Personalizado'. T ref vacío → 20 °C (T_N).  |  "
               "T_N = 20 °C = 293,15 K  |  "
               "1 scfm = 0,0283168 Nm³/min  |  "
               "FAD: caudal libre en condiciones de sala",
               C["text3"], ("Segoe UI", 9)).pack(anchor="w", pady=(4,0))

        btns = tk.Frame(body, bg=C["panel"])
        btns.pack(anchor="w", pady=(8, 2))

        tk.Button(btns, text="＋  Agregar equipo",
                  bg=C["card"], fg=C["accent"],
                  activebackground=C["sel"], activeforeground=C["accent"],
                  relief="flat", bd=0, padx=12, pady=5, cursor="hand2",
                  font=("Consolas", 9, "bold"),
                  highlightthickness=1, highlightbackground=C["border2"],
                  command=self._agregar_fila).pack(side="left", padx=(0,8))

        tk.Button(btns, text="✔  Confirmar equipos",
                  bg=C["ok"], fg=C["bg"],
                  activebackground="#00a040", activeforeground=C["bg"],
                  relief="flat", bd=0, padx=12, pady=5, cursor="hand2",
                  font=("Consolas", 9, "bold"),
                  command=self._confirmar_sec2).pack(side="left")

    def _confirmar_sec2(self):
        if not self.filas:
            self._sec2.set_estado("error", "SIN EQUIPOS")
            return
        try:
            for f in self.filas:
                f.obtener_datos()
        except ValueError as e:
            self._sec2.set_estado("error", "ERROR EN DATOS")
            messagebox.showerror("Error equipos", str(e))
            return
        n = len(self.filas)
        self._sec2.set_estado("ok", f"{n} EQUIPO{'S' if n>1 else ''}")
        self._sec2.cerrar()

    # ══════════════════════════════════════════
    #  SECCIÓN 3 – COMPRESOR
    # ══════════════════════════════════════════
    def _sec3_compresor(self):
        self._sec3 = SeccionPanel(self.main, "3", "Compresor requerido")
        body = self._sec3.body

        # ΔP cañerías
        r1 = tk.Frame(body, bg=C["panel"])
        r1.pack(fill="x", pady=(0,4))
        mk_lbl(r1, "ΔP pérdidas en cañerías:").pack(side="left", padx=(0,10))
        self.var_dp_can = tk.StringVar(value="fijo")
        mk_radio(r1, "0,1 bar fijo",    self.var_dp_can, "fijo").pack(side="left", padx=4)
        mk_radio(r1, "2% de P trabajo", self.var_dp_can, "pct" ).pack(side="left", padx=4)

        # ΔP sala
        r2 = tk.Frame(body, bg=C["panel"])
        r2.pack(fill="x", pady=(0,4))
        mk_lbl(r2, "ΔP pérdidas en sala de compresores (bar):").pack(side="left", padx=(0,8))
        self.var_dp_sala = tk.StringVar()
        mk_entry(r2, self.var_dp_sala, 8).pack(side="left")

        mk_sep(body)

        # Presión de trabajo
        r3 = tk.Frame(body, bg=C["panel"])
        r3.pack(fill="x", pady=(0,4))
        mk_lbl(r3, "Presión de trabajo:").pack(side="left", padx=(0,12))
        self.var_p_auto = tk.BooleanVar(value=True)
        tk.Checkbutton(r3, text="Usar P máxima de la tabla",
                       variable=self.var_p_auto,
                       command=self._toggle_p,
                       bg=C["panel"], fg=C["text"],
                       activebackground=C["panel"], activeforeground=C["accent"],
                       selectcolor=C["border"],
                       font=("Segoe UI", 10)).pack(side="left", padx=(0,16))
        mk_lbl(r3, "P manual (bar g):").pack(side="left", padx=(0,6))
        self.var_p_manual = tk.StringVar()
        self.entry_pm = mk_entry(r3, self.var_p_manual, 9)
        self.entry_pm.pack(side="left")
        mk_lbl(r3, "≥ P máx tabla", C["text3"], ("Segoe UI", 9)).pack(side="left", padx=8)
        self._toggle_p()

        mk_sep(body)

        # Factor de seguridad
        r4 = tk.Frame(body, bg=C["panel"])
        r4.pack(fill="x", pady=(0,4))
        mk_lbl(r4, "Factor de seguridad sobre caudal total (%):").pack(side="left", padx=(0,8))
        self.var_fs = tk.StringVar()
        mk_entry(r4, self.var_fs, 8).pack(side="left")
        mk_lbl(r4, "  Q compresor = Σ q_FAD_ef × (1 + FS/100)",
               C["text3"], ("Segoe UI", 9)).pack(side="left", padx=8)

        mk_sep(body)
        tk.Button(body, text="✔  Confirmar configuración del compresor",
                  bg=C["ok"], fg=C["bg"],
                  activebackground="#00a040", activeforeground=C["bg"],
                  relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
                  font=("Consolas", 9, "bold"),
                  command=self._confirmar_sec3).pack(anchor="w")

    def _toggle_p(self):
        auto = self.var_p_auto.get()
        self.entry_pm.config(
            state="disabled" if auto else "normal",
            highlightbackground=C["border"] if auto else C["border2"])

    def _confirmar_sec3(self):
        dp_s = self.var_dp_sala.get().strip()
        fs_s = self.var_fs.get().strip()
        if not self.var_p_auto.get():
            pm = self.var_p_manual.get().strip()
            if not pm:
                self._sec3.set_estado("error", "P MANUAL VACÍA")
                return
        self._sec3.set_estado("ok",
            f"ΔP sala={'?' if not dp_s else fmt(parse(dp_s),2)}  "
            f"FS={'0' if not fs_s else fmt(parse(fs_s),1)}%")
        self._sec3.cerrar()

    # ══════════════════════════════════════════
    #  SECCIÓN 4 – DEPÓSITO
    # ══════════════════════════════════════════
    def _sec4_deposito(self):
        self._sec4 = SeccionPanel(self.main, "4",
                                  "Depósito de aire comprimido",
                                  C["accent3"])
        body = self._sec4.body

        r1 = tk.Frame(body, bg=C["panel"])
        r1.pack(fill="x", pady=(0,6))
        mk_lbl(r1, "ΔP depósito (bar):").pack(side="left", padx=(0,6))
        dp_vals = [str(k).replace(".", ",") for k in sorted(delta_p_coefs)]
        self.var_dep_dp = tk.StringVar()
        ttk.Combobox(r1, textvariable=self.var_dep_dp,
                     values=dp_vals, state="readonly",
                     width=10, font=("Consolas", 10)).pack(side="left", padx=(0,28))
        mk_lbl(r1, "Conmutaciones Z (arranques/h):").pack(side="left", padx=(0,6))
        z_vals = [str(k) for k in sorted(z_coefs)]
        self.var_dep_z = tk.StringVar()
        ttk.Combobox(r1, textvariable=self.var_dep_z,
                     values=z_vals, state="readonly",
                     width=8, font=("Consolas", 10)).pack(side="left")

        r2 = tk.Frame(body, bg=C["panel"])
        r2.pack(fill="x", pady=(0,4))
        mk_lbl(r2, "Caudal base:").pack(side="left", padx=(0,10))
        self.var_dep_src = tk.StringVar(value="auto")
        mk_radio(r2, "Caudal calculado del compresor",
                 self.var_dep_src, "auto",
                 self._toggle_dep_q).pack(side="left", padx=(0,14))
        mk_radio(r2, "Manual (m³/min):",
                 self.var_dep_src, "manual",
                 self._toggle_dep_q).pack(side="left")
        self.var_dep_q_m = tk.StringVar()
        self.entry_dep_q = mk_entry(r2, self.var_dep_q_m, 9, state="disabled")
        self.entry_dep_q.pack(side="left", padx=6)

        mk_sep(body)
        tk.Button(body, text="✔  Confirmar configuración del depósito",
                  bg=C["ok"], fg=C["bg"],
                  activebackground="#00a040", activeforeground=C["bg"],
                  relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
                  font=("Consolas", 9, "bold"),
                  command=self._confirmar_sec4).pack(anchor="w")

    def _toggle_dep_q(self):
        manual = self.var_dep_src.get() == "manual"
        self.entry_dep_q.config(
            state="normal" if manual else "disabled",
            highlightbackground=C["border2"] if manual else C["border"])

    def _confirmar_sec4(self):
        dp = self.var_dep_dp.get().strip()
        z  = self.var_dep_z.get().strip()
        if not dp or not z:
            self._sec4.set_estado("error", "SELECCIONE ΔP y Z")
            return
        if self.var_dep_src.get() == "manual" and not self.var_dep_q_m.get().strip():
            self._sec4.set_estado("error", "CAUDAL MANUAL VACÍO")
            return
        self._sec4.set_estado("ok", f"ΔP={dp} bar  Z={z}/h")
        self._sec4.cerrar()

    # ── BOTÓN CALCULAR ──────────────────────────
    def _boton_calcular(self):
        f = tk.Frame(self.main, bg=C["bg"])
        f.pack(fill="x", pady=10)
        tk.Button(f, text="▶   CALCULAR",
                  bg=C["accent"], fg=C["bg"],
                  activebackground="#009acc", activeforeground=C["bg"],
                  relief="flat", bd=0, padx=36, pady=11, cursor="hand2",
                  font=("Consolas", 11, "bold"),
                  command=self._calcular).pack(anchor="center")

    # ══════════════════════════════════════════
    #  SECCIÓN 5 – RESULTADOS  (sin collapsar)
    # ══════════════════════════════════════════
    def _sec5_resultados(self):
        body = self._card("5 · Resultados", C["accent2"])

        # ── Layout: izquierda = tabla + resumen, derecha = diagrama ──
        split = tk.Frame(body, bg=C["panel"])
        split.pack(fill="x")

        # Panel izquierdo
        left = tk.Frame(split, bg=C["panel"])
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Panel derecho (diagrama)
        right = tk.Frame(split, bg=C["panel"])
        right.pack(side="right", fill="y")

        self.diagrama = Diagrama(right, width=380)
        self.diagrama.pack(fill="both", expand=True)

        # ── Tabla de caudales por equipo ──────
        tk.Label(left, text="CAUDALES POR EQUIPO",
                 bg=C["panel"], fg=C["text3"],
                 font=("Consolas", 8, "bold")).pack(anchor="w")

        RES_COLS = [
            (22, "Equipo"),
            (14, "q_N (Nl/min)"),
            (17, "q_FAD (m³/min)"),
            (7,  "Uso (%)"),
            (23, "q_FAD efectivo (m³/min)"),
        ]
        hdr_r = tk.Frame(left, bg=C["border"])
        hdr_r.pack(fill="x", pady=(4,0))
        for w, t in RES_COLS:
            tk.Label(hdr_r, text=t, bg=C["border"], fg=C["accent"],
                     font=("Consolas", 8, "bold"),
                     width=w, anchor="w").pack(side="left", padx=4, pady=3)

        self.frame_res = tk.Frame(left, bg=C["panel"])
        self.frame_res.pack(fill="x")

        # Total
        tot_f = tk.Frame(left, bg=C["card"],
                         highlightthickness=1,
                         highlightbackground=C["border2"])
        tot_f.pack(fill="x", pady=(2,8))
        tk.Label(tot_f, text="TOTAL  q_FAD efectivo",
                 bg=C["card"], fg=C["text2"],
                 font=("Consolas", 9, "bold")).pack(side="left", padx=12, pady=6)
        self._var_total = tk.StringVar(value="—")
        self.lbl_total = tk.Entry(tot_f, textvariable=self._var_total, width=22,
                                  bg=C["card"], readonlybackground=C["card"],
                                  fg=C["accent"], font=("Consolas", 11, "bold"),
                                  relief="flat", bd=0, highlightthickness=0,
                                  state="readonly")
        self.lbl_total.pack(side="right", padx=12)

        mk_sep(left)

        # ── Resumen compresor + depósito ──────
        cards = tk.Frame(left, bg=C["panel"])
        cards.pack(fill="x")

        # Compresor card
        cc = tk.Frame(cards, bg=C["card"],
                      highlightthickness=1, highlightbackground=C["border2"])
        cc.pack(side="left", fill="both", expand=True, padx=(0,4))
        tk.Label(cc, text="COMPRESOR",
                 bg=C["card"], fg=C["accent"],
                 font=("Consolas", 8, "bold")).pack(anchor="w", padx=10, pady=(8,4))

        self._rc = {}
        for key, lbl_txt, col in [
            ("q_comp",   "Caudal FAD (m³/min):",       C["accent"]),
            ("p_trab",   "Presión de trabajo (bar g):", C["accent2"]),
            ("p_base",   "P base (bar g):",             C["text2"]),
            ("dp_can",   "ΔP cañerías (bar):",          C["text2"]),
            ("dp_sala",  "ΔP sala (bar):",              C["text2"]),
            ("fs",       "Factor seguridad (%):",       C["text2"]),
        ]:
            rf = tk.Frame(cc, bg=C["card"])
            rf.pack(fill="x", padx=10, pady=2)
            tk.Label(rf, text=lbl_txt, bg=C["card"], fg=C["text3"],
                     font=("Segoe UI", 9)).pack(side="left")
            var = tk.StringVar(value="—")
            v = tk.Entry(rf, textvariable=var, width=22,
                         font=("Consolas", 10, "bold"), fg=col,
                         bg=C["card"], readonlybackground=C["card"],
                         relief="flat", bd=0, highlightthickness=0,
                         state="readonly")
            v.pack(side="right")
            v._var = var
            self._rc[key] = v

        # Depósito card
        dc = tk.Frame(cards, bg=C["card"],
                      highlightthickness=1, highlightbackground=C["border2"])
        dc.pack(side="left", fill="both", expand=True, padx=(4,0))
        tk.Label(dc, text="DEPÓSITO",
                 bg=C["card"], fg=C["accent3"],
                 font=("Consolas", 8, "bold")).pack(anchor="w", padx=10, pady=(8,4))

        self._rd = {}
        for key, lbl_txt, col in [
            ("vol",   "Volumen depósito (m³):",  C["accent3"]),
            ("q_dep", "Caudal base (m³/min):",   C["text2"]),
            ("dp",    "ΔP depósito (bar):",      C["text2"]),
            ("z",     "Conmutaciones Z (1/h):",  C["text2"]),
        ]:
            rf = tk.Frame(dc, bg=C["card"])
            rf.pack(fill="x", padx=10, pady=2)
            tk.Label(rf, text=lbl_txt, bg=C["card"], fg=C["text3"],
                     font=("Segoe UI", 9)).pack(side="left")
            var = tk.StringVar(value="—")
            v = tk.Entry(rf, textvariable=var, width=22,
                         font=("Consolas", 10, "bold"), fg=col,
                         bg=C["card"], readonlybackground=C["card"],
                         relief="flat", bd=0, highlightthickness=0,
                         state="readonly")
            v.pack(side="right")
            v._var = var
            self._rd[key] = v

        # Nota
        self.lbl_nota = tk.Label(body, text="",
                                 bg=C["panel"], fg=C["text3"],
                                 font=("Segoe UI", 8),
                                 wraplength=1100, justify="left")
        self.lbl_nota.pack(anchor="w", pady=(10,4))

    # ─── GESTIÓN FILAS ──────────────────────────
    def _agregar_fila(self):
        f = FilaEquipo(self.frame_tabla, on_delete=self._eliminar_fila)
        f.row.grid(row=self._next_eq_row, column=0,
                   columnspan=len(COL_DEFS), sticky="ew", pady=1)
        self._next_eq_row += 1
        self.filas.append(f)

    def _eliminar_fila(self, fila):
        if fila in self.filas:
            self.filas.remove(fila)
            fila.destroy()
            # Re-asignar grid rows para filas restantes (fila 0 = header)
            for i, f in enumerate(self.filas):
                f.row.grid(row=i+1, column=0,
                           columnspan=len(COL_DEFS), sticky="ew", pady=1)
                f._widgets["del"].config(
                    command=lambda x=f: self._eliminar_fila(x))
            self._next_eq_row = len(self.filas) + 1

    # ─── GUARDAR / CARGAR ───────────────────────
    def _guardar(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("Proyecto JSON", "*.json"), ("Todos", "*.*")],
            title="Guardar proyecto")
        if not path:
            return
        data = {
            "t_amb":      self.var_t_amb.get(),
            "p_atm":      self.var_p_atm.get(),
            "dp_can":     self.var_dp_can.get(),
            "dp_sala":    self.var_dp_sala.get(),
            "p_auto":     self.var_p_auto.get(),
            "p_manual":   self.var_p_manual.get(),
            "fs":         self.var_fs.get(),
            "dep_dp":     self.var_dep_dp.get(),
            "dep_z":      self.var_dep_z.get(),
            "dep_src":    self.var_dep_src.get(),
            "dep_q_m":    self.var_dep_q_m.get(),
            "equipos":    [f.to_dict() for f in self.filas],
        }
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
        messagebox.showinfo("Guardar", f"Proyecto guardado en:\n{path}")

    def _cargar(self):
        path = filedialog.askopenfilename(
            filetypes=[("Proyecto JSON", "*.json"), ("Todos", "*.*")],
            title="Cargar proyecto")
        if not path:
            return
        try:
            with open(path, encoding="utf-8") as fh:
                data = json.load(fh)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")
            return

        self.var_t_amb.set(data.get("t_amb", ""))
        self.var_p_atm.set(data.get("p_atm", ""))
        self.var_dp_can.set(data.get("dp_can", "fijo"))
        self.var_dp_sala.set(data.get("dp_sala", ""))
        self.var_p_auto.set(data.get("p_auto", True))
        self.var_p_manual.set(data.get("p_manual", ""))
        self.var_fs.set(data.get("fs", ""))
        self.var_dep_dp.set(data.get("dep_dp", ""))
        self.var_dep_z.set(data.get("dep_z", ""))
        self.var_dep_src.set(data.get("dep_src", "auto"))
        self.var_dep_q_m.set(data.get("dep_q_m", ""))

        # Limpiar y recargar filas
        for f in self.filas:
            f.destroy()
        self.filas.clear()
        for eq in data.get("equipos", []):
            f = FilaEquipo(self.frame_eq, on_delete=self._eliminar_fila)
            f.from_dict(eq)
            self.filas.append(f)

        self._toggle_p()
        self._toggle_dep_q()

        # Auto-confirmar secciones si tienen datos válidos
        self._confirmar_sec1()
        self._confirmar_sec2()
        self._confirmar_sec3()
        self._confirmar_sec4()

        messagebox.showinfo("Cargar", "Proyecto cargado correctamente.")

    # ─── CÁLCULO PRINCIPAL ──────────────────────
    def _calcular(self):
        try:
            # Condiciones ambientales
            t_s = self.var_t_amb.get().strip()
            p_s = self.var_p_atm.get().strip()
            if not t_s or not p_s:
                messagebox.showerror("Faltan datos",
                                     "Complete la sección de condiciones ambientales.")
                return
            t_amb = parse(t_s)
            p_atm = parse(p_s)

            dp_sala_s = self.var_dp_sala.get().strip()
            dp_sala   = parse(dp_sala_s) if dp_sala_s else 0.0
            fs_s      = self.var_fs.get().strip()
            fs        = parse(fs_s) if fs_s else 0.0

            if not self.filas:
                messagebox.showwarning("Sin equipos",
                                       "Agregue al menos un equipo en la sección 2.")
                return

            datos = [f.obtener_datos() for f in self.filas]

            total_FAD = 0.0
            max_p     = 0.0
            filas_r   = []

            for d in datos:
                if d["tipo"] == "FAD (m³/min)":
                    qFAD = d["caudal"] * d["cant"]
                    qN = qFAD * 1000.0  # solo para mostrar
                else:
                    qN = a_Nl_min(d["caudal"], d["tipo"], d["p_ref"] or 0.0, d["t_ref"], p_atm, t_amb_c=t_amb) * d["cant"]
                    qFAD = Nl_min_a_FAD(qN, t_amb, p_atm)
                qFAD_ef = qFAD * (d["uso"] / 100.0)
                total_FAD += qFAD_ef
                if d["p_req"] > max_p:
                    max_p = d["p_req"]
                filas_r.append((d["nombre"], d["cant"], qN, qFAD, d["uso"], qFAD_ef))
            # Guardar para exportar Excel
            self._last_filas_r = filas_r

            # Presión base
            if self.var_p_auto.get():
                p_base = max_p
            else:
                pm_s = self.var_p_manual.get().strip()
                if not pm_s:
                    messagebox.showerror("Presión manual",
                                         "Ingrese P manual o active 'Usar P máxima'.")
                    return
                p_base = parse(pm_s)
                if p_base < max_p:
                    messagebox.showerror("Error presión",
                        f"P manual ({fmt(p_base,2)} bar g) debe ser ≥ "
                        f"P máx tabla ({fmt(max_p,2)} bar g).")
                    return

            modo   = self.var_dp_can.get()
            dp_can = 0.1 if modo == "fijo" else p_base * 0.02
            p_trab = p_base + dp_can + dp_sala
            q_comp = total_FAD * (1.0 + fs / 100.0)

            # Depósito
            dp_v_s = self.var_dep_dp.get().strip()
            z_v_s  = self.var_dep_z.get().strip()
            if not dp_v_s or not z_v_s:
                messagebox.showerror("Depósito",
                                     "Seleccione ΔP y conmutaciones Z (sección 4).")
                return
            dp_v = parse(dp_v_s)
            z_v  = int(z_v_s)

            if self.var_dep_src.get() == "manual":
                qs = self.var_dep_q_m.get().strip()
                if not qs:
                    messagebox.showerror("Depósito",
                                         "Ingrese el caudal manual para el depósito.")
                    return
                q_dep = parse(qs)
            else:
                q_dep = q_comp

            tamano = calc_deposito(q_dep, dp_v, z_v)

            # ── Tabla de resultados ──
            for w in self.frame_res.winfo_children():
                w.destroy()

            for nombre, cant, qN, qFAD, uso, qFAD_ef in filas_r:
                rf = tk.Frame(self.frame_res, bg=C["panel"])
                rf.pack(fill="x", pady=1)
                for txt, cw, col in [
                    (f"{nombre} ×{cant}", 22, C["text"]),
                    (fmt(qN, 1),           14, C["text2"]),
                    (fmt(qFAD, 4),         17, C["text2"]),
                    (f"{uso:.0f}%",         7, C["text2"]),
                    (fmt(qFAD_ef, 4),      23, C["accent"]),
                ]:
                    var = tk.StringVar(value=txt)
                    e = tk.Entry(rf, textvariable=var, width=cw,
                                 font=("Consolas", 10), fg=col,
                                 bg=C["panel"], readonlybackground=C["panel"],
                                 relief="flat", bd=0,
                                 highlightthickness=0,
                                 state="readonly")
                    e.pack(side="left", padx=4)

            self._var_total.set(f"{fmt(total_FAD,4)} m³/min")

            # ── Cards compresor ──
            modo_txt = "fijo" if modo == "fijo" else "2% de P"
            self._rc["q_comp" ]._var.set(f"{fmt(q_comp,4)} m³/min")
            self._rc["p_trab" ]._var.set(f"{fmt(p_trab,3)} bar g")
            self._rc["p_base" ]._var.set(f"{fmt(p_base,2)} bar g")
            self._rc["dp_can" ]._var.set(f"{fmt(dp_can,4)} bar ({modo_txt})")
            self._rc["dp_sala"]._var.set(f"{fmt(dp_sala,3)} bar")
            self._rc["fs"     ]._var.set(f"{fmt(fs,1)} %")

            # ── Cards depósito ──
            if tamano is None:
                self._rd["vol"]._var.set("Caudal muy bajo para ΔP elegido")
                self._rd["vol"].config(fg=C["red"])
                v_str = "ERR"
            else:
                self._rd["vol"]._var.set(f"{fmt(tamano,4)} m³")
                self._rd["vol"].config(fg=C["accent3"])
                v_str = fmt(tamano, 3)

            self._rd["q_dep"]._var.set(f"{fmt(q_dep,4)} m³/min")
            self._rd["dp"   ]._var.set(f"{fmt(dp_v,2)} bar")
            self._rd["z"    ]._var.set(f"{z_v} 1/h")

            # Guardar todos los resultados calculados para Excel
            self._last_results = dict(
                t_amb=t_amb, p_atm=p_atm,
                total_FAD=total_FAD, q_comp=q_comp,
                p_base=p_base, dp_can=dp_can, dp_sala=dp_sala,
                p_trab=p_trab, fs=fs, modo=modo,
                dp_v=dp_v, z_v=z_v, q_dep=q_dep, tamano=tamano,
            )

            # ── Diagrama ──
            self.diagrama.update_values(
                q=fmt(q_comp, 3),
                p=fmt(p_trab, 2),
                v=v_str)

            self.lbl_nota.config(
                text=(
                    f"Referencia: T_N=0°C · P_N=1,01325 bar abs  |  "
                    f"FAD a T_amb={fmt(t_amb,1)}°C · P_atm={fmt(p_atm,5)} bar abs  |  "
                    f"FS={fmt(fs,1)}%  |  "
                    f"ΔP cañerías={fmt(dp_can,4)} bar  |  "
                    f"ΔP sala={fmt(dp_sala,3)} bar  |  "
                    f"scfm→Nl/min: ×0,0283168×(273,15/293,15)×1000 = ×26,3526  |  "
                    f"FAD→Nl/min: ×(P_atm/P_N)×(T_N/T_amb)×1000  |  "
                    f"Personalizado: ×(P_ref/P_N)×(T_N/T_ref)×1000"
                ))

        except ValueError as e:
            messagebox.showerror("Error de entrada", str(e))
        except Exception as e:
            messagebox.showerror("Error inesperado", str(e))



    # ─── EXPORTAR EXCEL ─────────────────────────
    def _exportar_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Excel no disponible",
                "No se encontró la librería openpyxl.\n"
                "Instálala con:  pip install openpyxl")
            return
        if not hasattr(self, "_last_results"):
            messagebox.showwarning("Sin resultados",
                "Primero calcule los resultados antes de exportar.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            title="Exportar a Excel")
        if not path:
            return

        r  = self._last_results
        fr = getattr(self, "_last_filas_r", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sala de Compresores"

        # Estilos
        def hdr_style(cell, color="1a3a5c"):
            cell.font      = Font(bold=True, color="FFFFFF", name="Consolas")
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def val_style(cell):
            cell.font      = Font(name="Consolas", size=10)
            cell.alignment = Alignment(horizontal="right")

        def sec_title(ws, row, text, cols=6):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=cols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(bold=True, color="00BFFF", name="Consolas", size=11)
            c.fill = PatternFill("solid", fgColor="111111")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20

        thin = Side(style="thin", color="333333")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row = 1
        # ── Título ──
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value="SALA DE COMPRESORES — Resultados")
        c.font = Font(bold=True, size=14, color="00BFFF", name="Consolas")
        c.fill = PatternFill("solid", fgColor="080808")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 28
        row += 2

        # ── Condiciones ambientales ──
        sec_title(ws, row, "1. CONDICIONES AMBIENTALES"); row += 1
        for label, val in [
            ("Temperatura ambiente (°C)",       r["t_amb"]),
            ("Presión atmosférica local (bar abs)", r["p_atm"]),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(name="Consolas")
            c = ws.cell(row=row, column=3, value=val)
            val_style(c); row += 1
        row += 1

        # ── Equipos / Servicios ──
        sec_title(ws, row, "2. SERVICIOS QUE REQUIEREN AIRE COMPRIMIDO"); row += 1
        hdrs2 = ["Equipo","Cant.","P req (bar g)","Caudal","Tipo caudal",
                 "Uso (%)","P ref (bar g)","T ref (°C)",
                 "q_N (Nl/min)","q_FAD (m³/min)","q_FAD efectivo (m³/min)"]
        for ci, h in enumerate(hdrs2, 1):
            c = ws.cell(row=row, column=ci, value=h)
            hdr_style(c); c.border = border
        row += 1
        for fila in self.filas:
            d = fila.to_dict()
            # Find matching result row
            match = next(((n,ca,qn,qf,us,qfe) for (n,ca,qn,qf,us,qfe) in fr
                          if n == (d.get("nombre","").strip() or "Equipo")), None)
            vals = [
                d.get("nombre",""), d.get("cant",""), d.get("p_req",""),
                d.get("caudal",""), d.get("tipo",""), d.get("uso",""),
                d.get("p_ref",""), d.get("t_ref",""),
                round(match[2],2) if match else "",
                round(match[3],4) if match else "",
                round(match[5],4) if match else "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = Font(name="Consolas", size=10)
                c.border = border
            row += 1

        # Total
        ws.cell(row=row, column=9,  value="TOTAL q_FAD efectivo").font = Font(bold=True, name="Consolas")
        c = ws.cell(row=row, column=11, value=round(r["total_FAD"],4))
        c.font = Font(bold=True, name="Consolas", color="00BFFF")
        row += 2

        # ── Compresor ──
        sec_title(ws, row, "3. COMPRESOR REQUERIDO"); row += 1
        modo_txt = "fijo" if r["modo"] == "fijo" else "2% de P"
        for label, val in [
            ("Caudal FAD compresor (m³/min)",    round(r["q_comp"],4)),
            ("Presión de trabajo (bar g)",        round(r["p_trab"],3)),
            ("P base equipos (bar g)",            round(r["p_base"],2)),
            (f"ΔP cañerías ({modo_txt}) (bar)",  round(r["dp_can"],4)),
            ("ΔP sala (bar)",                     round(r["dp_sala"],3)),
            ("Factor de seguridad (%)",           round(r["fs"],1)),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(name="Consolas")
            c = ws.cell(row=row, column=3, value=val)
            val_style(c); row += 1
        row += 1

        # ── Depósito ──
        sec_title(ws, row, "4. DEPÓSITO DE AIRE COMPRIMIDO"); row += 1
        tam = r["tamano"]
        for label, val in [
            ("Volumen depósito (m³)",  round(tam,4) if tam else "ERR"),
            ("Caudal base (m³/min)",   round(r["q_dep"],4)),
            ("ΔP depósito (bar)",      round(r["dp_v"],2)),
            ("Conmutaciones Z (1/h)",  r["z_v"]),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(name="Consolas")
            c = ws.cell(row=row, column=3, value=val)
            val_style(c); row += 1

        # Ajustar anchos de columna
        col_widths = [28, 8, 16, 14, 22, 8, 12, 10, 16, 16, 24]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        try:
            wb.save(path)
            messagebox.showinfo("Excel exportado", f"Archivo guardado en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))

# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()